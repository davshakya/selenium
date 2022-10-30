import openpyxl

workbook=openpyxl.load_workbook("D:\\demo\\selenium\\pytest_selenium_practice\\dev.xlsx")
sheet=workbook["Sheet"]
a=sheet.cell(1,2).value
print(a)
#
# workbook=openpyxl.Workbook()
# sheet=workbook.active
# sheet.cell(1,1).value="Hello dev"
# workbook.save("D:\\demo\\pytest_selenium_practice\\dev.xlsx")
#
# f=open("data_file.txt","a+")
# f.write("Hello Shakya")
# f.write("How are u")
# f.close()
#
# from PIL import Image
# image = Image.open("C:\\work\\test_img.jpg")
# image = image.resize((500,500),Image.ANTIALIAS)
# image = image.resize((500,500),Image.ANTIALIAS)
# image.save(fp="newimage.jpg")
# import openpyxl

# l = [3, 4, 5, 6, 1, 2]
# x = list(map(lambda a: a * 2, l))
# print(x)
#
# y = list(filter(lambda a: a > 2, l))
# print(y)
#
# z = reduce(lambda a, b: a + b, l)
# print(z)
#
# list_comp = [item * 2 for item in l]
# print(list_comp)
#
# list_comp1 = [item * 2 for item in l]
# print(list_comp)
#
# dt = {"e": 3, "b": 5, "c": 9, "d": 6}
# x=sorted(dt.values())
# print(x)
# # d = {value: value ** 2 for value in l}
# # print(d)
# d = {key: value for key, value in sorted(dt.items())}
# print(d)
# print(d["e"])


# a="devendra shakya"
# b="Ram123"
# c="Kumar"

# import mysql.connector as ms
#
# mydb=ms.connect(host="localhost",
#                 port=3306,
#                 user="root",
#                 database="",
#                 )
# if mydb.is_connected:

# wb=openpyxl.load_workbook("C:\\work\\test_exl.xlsx")
# sheet=wb["Sheet1"]
# for i in range(1,5):
#     for j in range(1,3):
#         x=sheet.cell(i,j).value
#         print(x)

# wbk=openpyxl.Workbook()
# sheet=wbk.active
# l=[(12,5,4),(4,4),(5,3),(32,5)]
# sheet.cell(1,2).value="shakya"
# sheet.cell(6,1).value="Devendra"
# for i in l:
#     sheet.append(i)
# wbk.save("dev.xlsx")


#
#
# income = [('Income', ''),
#           ('Salary', 1000),
#           ('Investment', 500),
#           ('Side hustle', 500),
#           ('aTotal', 2000)]
#
# expense = [('Expense',''),
#            ('Housing', 1200),
#            ('Insurance', 200),
#            ('Grocery', 500),
#            ('Entertainment', 500),
#            ('Total', 2400)]
#
# for row in income+expense:
#     ws.append(row)
#
# wb.save('formatting.xlsx')
#
#


# print(d["b"][0])
# print(len(d["b"]))


# for k in range(len(d["b"])):

#     d["b"][k]=11

# print(d)


# try:
#     a=int(input("Enter any value of A ="))
#     b=int(input("Enter any value of B ="))
#     c=a/b
#     print(c)
# except ZeroDivisionError:
#     print("ZeroDivisionError exception entered")
# except ValueError:
#     print("ValueError exception came")
# finally:
#     print("finally block run always")
#
#
# a = "dev"
# b = "shakya"
# print(f"hello {a},{b}")
#
# try:
#     a = 3 / 0
# except:
#     print("invalid")
